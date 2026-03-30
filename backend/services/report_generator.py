"""Generate XLSX validation reports using openpyxl.

Creates a detailed report with transaction-level and credit-level validation results.
"""
import json
import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def generate_report(test_run, transactions, output_dir):
    """Generate an XLSX validation report.

    Args:
        test_run: Test run dict from DB.
        transactions: List of transaction dicts from DB.
        output_dir: Directory to write the report file.

    Returns:
        Path to the generated report file.
    """
    run_id = test_run.get("id")
    logger.info("Generating report for test run id=%d (%d transactions)",
                run_id, len(transactions))

    wb = Workbook()

    # ── Summary Sheet ────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, test_run, transactions)

    # ── Transaction Details Sheet ────────────────────────────────────────
    ws_txn = wb.create_sheet("Transactions")
    _build_transactions_sheet(ws_txn, transactions)

    # ── Credit Details Sheet ─────────────────────────────────────────────
    ws_credits = wb.create_sheet("Credit Details")
    _build_credits_sheet(ws_credits, transactions)

    # Save report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_run{run_id}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    logger.info("Report generated: %s", filepath)
    return filepath


def _build_summary_sheet(ws, test_run, transactions):
    """Build the summary sheet with overall stats."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # Title
    ws.append(["XML Payment Test Automation - Validation Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")
    ws.append([])

    # Test run info
    info = [
        ("Test Run ID", test_run.get("id")),
        ("Scheme", test_run.get("scheme_name", "N/A")),
        ("Upload File", test_run.get("upload_filename")),
        ("Created At", test_run.get("created_at")),
        ("Status", test_run.get("status")),
        ("Total Transactions", test_run.get("total_transactions")),
    ]
    for label, value in info:
        ws.append([label, str(value) if value else ""])

    ws.append([])

    # Validation summary
    total = len(transactions)
    passed = sum(1 for t in transactions
                 if _get_validation_overall(t) == "PASS")
    failed = sum(1 for t in transactions
                 if _get_validation_overall(t) == "FAIL")
    pending = total - passed - failed

    ws.append(["Validation Summary"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    ws.append(["Total", total])
    ws.append(["Passed", passed])
    ws.append(["Failed", failed])
    ws.append(["Pending/Not Validated", pending])

    # Color the results
    for row in ws.iter_rows(min_row=ws.max_row - 2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER


def _build_transactions_sheet(ws, transactions):
    """Build the transactions detail sheet."""
    headers = [
        "TC ID", "Batch Reference", "Debit Account", "Debit Account Parent", "Debit IFSC",
        "Debit Amount", "Credit Count", "Expected Status",
        "Actual Debit Status", "Actual Debit Remarks",
        "Initiation Validation", "Initiation Description",
        "Response Validation", "Response Description",
        "Overall Result", "Status"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Write data
    for row_idx, txn in enumerate(transactions, 2):
        validation_overall = _get_validation_overall(txn)
        row_data = [
            txn.get("tc_id"),
            txn.get("batch_reference"),
            txn.get("debit_account"),
            txn.get("debit_account_parent"),
            txn.get("debit_ifsc"),
            txn.get("debit_amount"),
            txn.get("credit_count"),
            txn.get("expected_status"),
            txn.get("actual_debit_status"),
            txn.get("actual_debit_remarks"),
            txn.get("initiation_validation", ""),
            txn.get("initiation_validation_desc", ""),
            txn.get("response_validation", ""),
            txn.get("response_validation_desc", ""),
            validation_overall,
            txn.get("status"),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(value) if value else "")
            cell.border = THIN_BORDER

            # Color validation result columns (init=11, resp=13, overall=15)
            if col in (11, 13, 15):
                val = str(value).upper() if value else ""
                if val == "PASS":
                    cell.fill = PASS_FILL
                elif val == "FAIL":
                    cell.fill = FAIL_FILL
                elif val:
                    cell.fill = PENDING_FILL

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18


def _build_credits_sheet(ws, transactions):
    """Build the credit-level details sheet."""
    headers = [
        "TC ID", "Batch Reference", "Credit #", "Account", "IFSC",
        "Expected Amount", "Credit Reference",
        "Initiation Status", "Initiation Remarks",
        "Response Status", "Response Remarks", "Response Amount",
        "Response Fail Status", "Response Fail Remarks",
        "Unique Credit Resp ID", "Validation"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row_idx = 2
    for txn in transactions:
        credits = txn.get("credit_json")
        if isinstance(credits, str):
            try:
                credits = json.loads(credits)
            except (json.JSONDecodeError, TypeError):
                credits = []
        credits = credits or []

        for c_idx, credit in enumerate(credits):
            row_data = [
                txn.get("tc_id"),
                txn.get("batch_reference"),
                c_idx + 1,
                credit.get("account"),
                credit.get("ifsc"),
                credit.get("amount"),
                credit.get("credit_reference"),
                credit.get("initiation_status"),
                credit.get("initiation_remarks"),
                credit.get("response_status"),
                credit.get("response_remarks"),
                credit.get("response_amount"),
                credit.get("response_fail_status"),
                credit.get("response_fail_remarks"),
                credit.get("unique_credit_resp_id"),
                credit.get("validation_result"),
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=str(value) if value else "")
                cell.border = THIN_BORDER

                if col == 16:  # Validation column
                    vr = str(value).upper() if value else ""
                    if vr == "PASS":
                        cell.fill = PASS_FILL
                    elif vr == "FAIL":
                        cell.fill = FAIL_FILL

            row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18


def _get_validation_overall(txn):
    """Extract overall validation result from a transaction."""
    vr = txn.get("validation_result")
    if isinstance(vr, str):
        try:
            vr = json.loads(vr)
        except (json.JSONDecodeError, TypeError):
            return vr or ""
    if isinstance(vr, dict):
        return vr.get("overall", "")
    return ""
