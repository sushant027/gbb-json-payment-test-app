"""Parse uploaded Excel files containing test data.

Handles pipe-delimited multi-credit fields (credit_account, credit_ifsc, etc.).
Each row becomes one transaction with its own credit_json array.
"""
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Known Excel column names (lowercase, stripped)
EXPECTED_COLUMNS = [
    "tcid", "scheme", "debit_account", "debit_account_parent", "debit_ifsc", "debit_amount",
    "credit_account", "credit_ifsc", "credit_count", "credit_amount",
    "beneficiary_name", "pay_mode", "expected_result"
]

# Columns that are pipe-delimited for multi-credit
PIPE_DELIMITED_COLUMNS = [
    "credit_account", "credit_ifsc", "credit_amount",
    "beneficiary_name", "pay_mode"
]


def parse_excel(file_path):
    """Parse an Excel file and return a list of transaction dicts.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        List of dicts, each representing one transaction row:
        {
            "tc_id": "TC11",
            "scheme": "MAHARASHTRA",
            "debit_account": "50117072025411",
            "debit_ifsc": "HDFC0000051",
            "debit_amount": 50000.0,
            "credit_count": 5,
            "expected_status": "SUCCESS",
            "credits": [
                {"account": "...", "ifsc": "...", "amount": "...",
                 "beneficiary_name": "...", "pay_mode": "..."},
                ...
            ]
        }
    """
    logger.info("Parsing Excel file: %s", file_path)

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    logger.debug("Active sheet: %s", ws.title)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.warning("Excel file is empty")
        return []

    # Parse header row - normalize to lowercase, strip whitespace
    raw_headers = rows[0]
    headers = []
    for h in raw_headers:
        if h is None:
            headers.append("")
        else:
            headers.append(str(h).strip().lower().replace(" ", "_"))

    logger.debug("Parsed headers: %s", headers)

    transactions = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if all(cell is None for cell in row):
            logger.debug("Skipping empty row %d", row_idx)
            continue

        # Build row dict
        row_data = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val is not None:
                val = str(val).strip()
            row_data[header] = val

        logger.debug("Processing row %d: tc_id=%s", row_idx, row_data.get("tcid"))

        # Handle debit_account that may come as scientific notation from Excel
        debit_account = row_data.get("debit_account", "")
        if debit_account:
            try:
                debit_account = str(int(float(debit_account)))
            except (ValueError, TypeError):
                debit_account = str(debit_account)

        # Handle debit_account_parent that may come as scientific notation from Excel
        debit_account_parent = row_data.get("debit_account_parent", "")
        if debit_account_parent:
            try:
                debit_account_parent = str(int(float(debit_account_parent)))
            except (ValueError, TypeError):
                debit_account_parent = str(debit_account_parent)

        # Parse debit amount
        try:
            debit_amount = float(row_data.get("debit_amount", 0))
        except (ValueError, TypeError):
            debit_amount = 0.0
            logger.warning("Row %d: invalid debit_amount, defaulting to 0", row_idx)

        # Parse credit count
        try:
            credit_count = int(float(row_data.get("credit_count", 0)))
        except (ValueError, TypeError):
            credit_count = 0

        # Parse pipe-delimited credit fields
        credit_accounts = _split_pipe(row_data.get("credit_account", ""))
        credit_ifscs = _split_pipe(row_data.get("credit_ifsc", ""))
        credit_amounts = _split_pipe(row_data.get("credit_amount", ""))
        beneficiary_names = _split_pipe(row_data.get("beneficiary_name", ""))
        pay_modes = _split_pipe(row_data.get("pay_mode", ""))

        # Determine actual credit count from data
        actual_count = max(
            len(credit_accounts), len(credit_ifscs),
            len(credit_amounts), credit_count
        )

        if credit_count and actual_count != credit_count:
            logger.warning("Row %d: credit_count=%d but found %d credit entries",
                           row_idx, credit_count, actual_count)

        # Build credits array
        credits = []
        for i in range(actual_count):
            credit = {
                "account": _safe_get(credit_accounts, i, ""),
                "ifsc": _safe_get(credit_ifscs, i, ""),
                "amount": _safe_get(credit_amounts, i, "0"),
                "beneficiary_name": _safe_get(beneficiary_names, i, ""),
                "pay_mode": _safe_get(pay_modes, i, ""),
                "credit_reference": "",
                "initiation_status": "",
                "initiation_remarks": "",
                "response_status": "",
                "response_remarks": "",
                "response_amount": "",
                "validation_result": ""
            }
            credits.append(credit)

        txn = {
            "tc_id": row_data.get("tcid", f"TC_{row_idx}"),
            "scheme": row_data.get("scheme", ""),
            "debit_account": debit_account,
            "debit_account_parent": debit_account_parent,
            "debit_ifsc": row_data.get("debit_ifsc", ""),
            "debit_amount": debit_amount,
            "credit_count": actual_count,
            "expected_status": row_data.get("expected_result", ""),
            "credits": credits
        }
        transactions.append(txn)
        logger.debug("Row %d parsed: tc_id=%s, credits=%d",
                      row_idx, txn["tc_id"], len(credits))

    wb.close()
    logger.info("Parsed %d transactions from Excel", len(transactions))
    return transactions


def _split_pipe(value):
    """Split a pipe-delimited string into a list."""
    if not value:
        return []
    return [v.strip() for v in str(value).split("|")]


def _safe_get(lst, idx, default=""):
    """Safely get an item from a list by index."""
    return lst[idx] if idx < len(lst) else default
